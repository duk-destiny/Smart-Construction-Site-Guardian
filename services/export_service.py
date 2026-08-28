"""报告导出服务（M08）：将工单/风险/任务汇总导出为 Excel 台账。

依赖：openpyxl。审计通过 AuditService（仅 INSERT）记录。
"""
from __future__ import annotations

import os
import sqlite3
from datetime import datetime

from dao.db import get_conn
from core.paths import data_path
from services.permission_service import PermissionService

_COLUMNS = ["工单ID", "任务ID", "隐患描述", "违反规范", "整改要求",
            "风险等级", "工人提示", "生成时间"]


class ExportService:
    """工单/风险台账导出。"""

    def __init__(self, conn: sqlite3.Connection | None = None):
        self._conn = conn
        self._permissions = PermissionService(self._get_conn())

    def _get_conn(self) -> sqlite3.Connection:
        if self._conn is not None:
            return self._conn
        return get_conn()

    def export_excel(self, task_id: str | None = None,
                    date_from: str | None = None,
                    date_to: str | None = None,
                    user_id: str | None = None) -> dict:
        """导出 Excel，返回 {ok, data:{file_path}}。"""
        if user_id:
            self._permissions.require(user_id, "export")
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        conn = self._get_conn()
        sql = (
            "SELECT w.id, w.task_id, w.hazard_desc, w.clause, w.requirement, "
            "w.risk_level, w.worker_notice, w.created_at "
            "FROM work_orders w"
        )
        clauses, params = [], []
        if task_id:
            clauses.append("w.task_id=?")
            params.append(task_id)
        if date_from:
            clauses.append("w.created_at>=?")
            params.append(date_from)
        if date_to:
            clauses.append("w.created_at<=?")
            params.append(date_to)
        if clauses:
            sql += " WHERE " + " AND ".join(clauses)
        sql += " ORDER BY w.created_at DESC"

        rows = conn.execute(sql, params).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "动火安全台账"
        ws.append(_COLUMNS)
        header_fill = PatternFill("solid", fgColor="C00000")
        for col, _ in enumerate(_COLUMNS, 1):
            c = ws.cell(row=1, column=col)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
        for r in rows:
            ws.append(list(r))
        ws.freeze_panes = "A2"
        for col in range(1, len(_COLUMNS) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

        os.makedirs(data_path("exports"), exist_ok=True)
        fname = f"动火安全台账_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        fpath = os.path.join(data_path("exports"), fname)
        wb.save(fpath)

        if self._conn is None:
            conn.close()
        return {"ok": True, "data": {"file_path": fpath, "rows": len(rows)}}


def load_export_file(name: str) -> tuple[str, str]:
    """校验导出文件名并解析为 (绝对路径, 展示名)，供 API 下载端点使用。

    下载防护：只接受纯文件名（basename 后必须与输入一致），解析结果必须
    落在 data/exports 目录内——阻断 ../ 等路径穿越；不存在抛 FileNotFoundError。
    """
    safe = os.path.basename((name or "").strip())
    if not safe or safe in (".", ".."):
        raise ValueError("非法的导出文件名")
    base = os.path.abspath(data_path("exports"))
    path = os.path.abspath(os.path.join(base, safe))
    if os.path.dirname(path) != base:
        raise ValueError("非法的导出文件名")
    if not os.path.isfile(path):
        raise FileNotFoundError(safe)
    return path, safe
